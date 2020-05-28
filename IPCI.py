# -*- coding: utf8 -*-
import os

#Import library
import xlrd 
import datetime
from openpyxl import Workbook

from tkinter import *
from tkinter import filedialog



def choix_doc():
    master.filename =  filedialog.askopenfilename(initialdir = "/",title = "Select file",filetypes = (("excel files","*.xls"),("all files","*.*")))
    file=master.filename.split("/")[-1]
        
    label_2.delete(0, END)
    label_2.insert(0, file)
def quitter():
    sys.exit()

def traitement():
    try:
            #changement de repectoire 

        wb= xlrd.open_workbook(master.filename)
        name_sheet=wb.sheet_names()[0] 
        
        # makke list
        ma_liste=[]
        sh = wb.sheet_by_name(name_sheet)
        for rownum in range(sh.nrows):
                ma_liste.append(sh.row_values(rownum))
            
        # new list pour ceux dont la date de derniere visite et la date de derniere dispensation diffère
        
        index_prop=7
        list_prb=[]
        list_clean=ma_liste[index_prop:]
        list_clean1=ma_liste[index_prop:]
        for i in list_clean:
            if i[11]!=i[13]:
                list_prb.append(i)
                
                
        #traitement de la date (le nombre de jours de TARV)
        
        list_date=[]
        list_10=[]
        for i in list_clean1:
            date=xlrd.xldate.xldate_as_datetime(int(i[11]), wb.datemode)
            date2=xlrd.xldate.xldate_as_datetime(int(i[14]), wb.datemode)
            date3=date + datetime.timedelta(i[12])
            date31=date3.strftime("%d %b %Y")
            if date2>date3: #date de prochain rdv sup ou egal a la date disp + nbre de jrs disp
                i.append(date31)
                list_date.append(i)
                #if int(str(date2-date3).split(' ')[0])<0:
                i.append(int(str(date2-date3).split(' ')[0]))
                list_10.append(i)
                    
        
        #traitement de la date( ceux dont la date de derniere visiste et la date de derniere dispensation)
        
        for i in list_prb:
            date=xlrd.xldate.xldate_as_datetime(int(i[11]), wb.datemode)
            date2=xlrd.xldate.xldate_as_datetime(int(i[13]), wb.datemode)
            i[11]=date.strftime("%d %b %Y")
            i[13]=date2.strftime("%d %b %Y")
            
        # second traitement TARV
        
        for i in list_date:
            try:
                date=xlrd.xldate.xldate_as_datetime(int(i[11]), wb.datemode)
                date2=xlrd.xldate.xldate_as_datetime(int(i[13]), wb.datemode)
                i[11]=date.strftime("%d %b %Y")
                i[13]=date2.strftime("%d %b %Y")
            except:
                date=i[11]
                date2=i[13]
            date14=xlrd.xldate.xldate_as_datetime(int(i[14]), wb.datemode)
            i[14]=date14.strftime("%d %b %Y")
            
            
        # Creation de fichier excel 
        
        from openpyxl.styles import NamedStyle, Font, Border, Side, alignment, Alignment
        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True, size=12)
        bd = Side(style='thick', color="000000")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        highlight.alignment = Alignment(horizontal="center", vertical="center")
        
        #creation du dossier de reception
        path=r"‪C:\Users\User\Documents\Projet_IPCI"
        os.chdir(path[1:])
        
        wbc = Workbook()
        sheet = wbc.active
        sheet.title = 'Liste patients PDV de 10 jours'
        wbc.add_named_style(highlight)
        for i in range(1,len(list_10)+3):
            sheet.merge_cells(F"A{i}:C{i}")
            sheet.merge_cells(F"D{i}:G{i}")
            sheet.merge_cells(F"H{i}:K{i}")
            sheet.merge_cells(F"L{i}:N{i}")
            sheet.merge_cells(F"O{i}:R{i}")
            sheet.merge_cells(F"S{i}:U{i}")
        # Print the titles into Excel Workbook:
        row = 1
        sheet['A'+str(row)] = ma_liste[6][1]
        sheet['D'+str(row)] = ma_liste[6][11]
        sheet['H'+str(row)] = "Nbre de Jours dispensés SIGDEP"
        sheet['L'+str(row)] = ma_liste[6][14]+"SIGDEP"
        sheet['O'+str(row)] = "Date de Prochain RDV calculé"
        sheet['S'+str(row)] = "PDV en nombre de Jours"
        for i,j in enumerate(list_10):
            sheet['A'+str(i+2)] = j[1]
            sheet['D'+str(i+2)] = j[11]
            sheet['H'+str(i+2)] = j[12]
            sheet['L'+str(i+2)] = j[14]
            sheet['O'+str(i+2)] = j[-2]
            sheet['S'+str(i+2)] = j[-1]
        for i in range(1,len(list_10)+3):
            for j in ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U"]:
                sheet[F"{j}{i}"].style = highlight
        sheet.sheet_view.showGridLines = False
        sheet.protection.sheet = True
        
        sheet1=wbc.create_sheet()
        #sheet1 = wbc.active
        sheet1.title = 'Liste Dates differentes'

        # Print the titles into Excel Workbook:
        for i in range(1,len(list_prb)+3):
            sheet1.merge_cells(F"A{i}:C{i}")
            sheet1.merge_cells(F"D{i}:G{i}")
            sheet1.merge_cells(F"H{i}:J{i}")
        row = 1
        sheet1['A'+str(row)] = ma_liste[6][1]
        sheet1['D'+str(row)] = ma_liste[6][11]
        sheet1['H'+str(row)] = ma_liste[6][13]
        for i,j in enumerate(list_prb):
            sheet1['A'+str(i+2)] = j[1]
            sheet1['D'+str(i+2)] = j[11]
            sheet1['H'+str(i+2)] = j[13]
        for i in range(1,len(list_prb)+3):
            for j in ["A","B","C","D","E","F","G","H","I","J"]:
                sheet1[F"{j}{i}"].style = highlight
        sheet1.sheet_view.showGridLines = False
        sheet1.protection.sheet = True

        # Save a file by date:
        filename = 'Patients Futurs PDV'+ datetime.datetime.now().strftime(" le %Y_%m_%d à %I H %M min_%S sec") + '.xlsx'
        wbc.save(filename)
        global folder_b
        folder_b.destroy()
        folder_b=Button(fr_s,text='Traitement en cours, Excel va bientôt demarrer',font=("algerian", 25),bg='#1ADEC0',fg='white',command=traitement)
        folder_b.pack(fill=X)
        

        # Open the file for the user:

        os.system('start excel.exe "%s\\%s"' % (path[1:], filename, ))
        master.destroy()
    except AttributeError:
         global label_2
         label_2.destroy()
         label_2=Entry(fr_s,font=("algerian", 25),bg='white',fg='#1ADEC0')
         label_2.pack() #envoi dans l'affichage
         label_2.delete(0, END)
         label_2.insert(0, "SELECTIONNER UN FICHIER")


#creat windows

master = Tk() # fenetre
master.title("Traitement IPCI") #title
master.geometry("700x360") # defaul_size 
master.minsize(600, 360) # min size interface
master.iconbitmap('dcdj.ico') # image logo
master.config(background='#1ADEC0')

#creer un frame
fr=Frame(master,bg='#1ADEC0')
fr_s=Frame(fr,bg='#1ADEC0')


#Creation image:
width=300
height=100
image=PhotoImage(file="dcdj.png").zoom(23).subsample(34)
canvas=Canvas(fr,width=width,height=height,bg='#1ADEC0',bd=1,highlightthickness=0)
canvas.create_image(width/2,height/2,image=image)
canvas.grid(row=0,column=0,sticky=N)


#text
label_1=Label(fr_s,text="Le Ficher chargé",font=("algerian", 25),bg='#1ADEC0',fg='white')
label_1.pack() #envoi dans l'affichage

label_2=Entry(fr_s,font=("algerian", 25),bg='#1ADEC0',fg='white')
label_2.pack() #envoi dans l'affichage

#first button
folder_b=Button(fr_s,text='Traiter le fichier',font=("algerian", 25),bg='#1ADEC0',fg='white',command=traitement)
folder_b.pack(fill=X)

fr_s.grid(row=1,column=0,sticky=S)



#creer un menu

menu_bar=Menu(master)
file_menu=Menu(menu_bar, tearoff=0)
file_menu.add_command(label="Fichier",command=choix_doc)
file_menu.add_command(label="Quitter",command=master.destroy)
menu_bar.add_cascade(label="MENU", menu=file_menu)

master.config(menu=menu_bar)

fr.pack(expand=YES)
master.mainloop()
os.system("pause")








